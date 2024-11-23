from pandas import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


model = "G+10"
typefile = "Stiffness"
displX = read_excel(f'Analysis/ExcelComparison/outputXYG+10Stiffness.xlsx')


def TableFormatter (displX,model, typefile):
    displX_new = displX.drop(columns=['Elevation(m)', 'Location'])
    displX_new1 = displX_new.drop(columns=['Unnamed: 0','Elevation(m).1', 'Location.1'])
    displX_rename = displX_new1.rename (columns={
        "X-Dir(conc)": "X-Dir", 
        "Y-Dir(conc)": "Y-Dir",
        "X-Dir(cnt)":"X-Dir",
        "Y-Dir(cnt)":"Y-Dir",
        "Percentage_Change in X":"%Change_XDir",
        "Percentage_Change in Y":"%Change_YDir"
    })
    
    displX_rename.insert(0,"Concrete", " ")
    displX_rename.insert(4,"CNT", " ")
    displX_rename.insert(8," ", " ")
    
    donwload_path = f'D:/Milset Asia Python/Analysis/ExcelComparison/outputXY{model}{typefile}.xlsx'
    
    with ExcelWriter(donwload_path) as writer:
        displX_rename.to_excel(writer)
    
    wb_style = load_workbook(f'D:/Milset Asia Python/Analysis/ExcelComparison/outputXY{model}{typefile}.xlsx')
    sheet = wb_style.active
    for rows in sheet.iter_cols(min_col=2, max_col=5, min_row=1, max_row=None):
        for cell in rows:
            cell.fill = PatternFill(start_color="808080", end_color="808080",fill_type = "solid")
    
    
    for rows in sheet.iter_cols(min_col=6, max_col=9, min_row=1, max_row=None):
        for cell in rows:
            cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6",fill_type = "solid")
    
    wb_style.save(f'D:/Milset Asia Python/Analysis/ExcelComparison/FormattedXY{model}{typefile}.xlsx')


TableFormatter(displX, model, typefile)